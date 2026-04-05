"""
xlsx_parser.py - Parses school lesson plan xlsx files into Python dicts
ready to be inserted into the database.
"""
import re
import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_xlsx(filepath: str) -> dict:
    """Parse a lesson plan xlsx file and return a structured dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the chapter sheet (case-insensitive)
    chapter_ws = None
    for name in wb.sheetnames:
        if name.lower() == "chapter":
            chapter_ws = wb[name]
            break
    if chapter_ws is None:
        raise ValueError(
            f"No 'Chapter' sheet found in {filepath}. Sheets: {wb.sheetnames}"
        )

    title, aim, concepts = _parse_chapter_sheet(chapter_ws)

    # Parse each exhibit sheet — determine exhibit key from content or sheet name
    exhibits: dict = {}
    for name in wb.sheetnames:
        if name.lower() == "chapter":
            continue
        ws = wb[name]

        # Prefer exhibit number from row 1 cell content (more reliable than sheet name)
        content_key = ""
        first_rows = list(ws.iter_rows(values_only=True, min_row=1, max_row=1))
        if first_rows:
            for cell_val in first_rows[0]:
                if cell_val is not None:
                    candidate = _normalize_exhibit_ref(str(cell_val))
                    if candidate:
                        content_key = candidate
                        break

        # Fall back to sheet name if content didn't yield a key
        normalized_key = content_key or _normalize_exhibit_ref(name)

        if normalized_key:
            exhibit_data = _parse_exhibit_sheet(ws)
            exhibits[normalized_key] = exhibit_data

    return {
        "title": title,
        "aim": aim,
        "concepts": concepts,
        "exhibits": exhibits,
    }


# ---------------------------------------------------------------------------
# Chapter sheet parser
# ---------------------------------------------------------------------------

def _parse_chapter_sheet(ws) -> tuple:
    """
    Parse the Chapter sheet.
    Returns (title, aim, concepts).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Row 1: title — may be in col 0 or col 1 (Hindi has it in col 0)
    title_row = rows[0] if rows else ()
    title = _clean(title_row[0] if title_row else None)
    if not title and len(title_row) > 1:
        title = _clean(title_row[1])

    # Row 2: aim — may be in col 0 (most) or col 1 (Hindi)
    aim_row = rows[1] if len(rows) > 1 else ()
    aim_raw = _clean(aim_row[0] if aim_row else None)
    if not aim_raw and len(aim_row) > 1:
        aim_raw = _clean(aim_row[1])
    aim = _strip_aim_prefix(aim_raw)

    # Rows 5+ (index 4+): concept rows
    # Row 3 (index 2) = column headers, Row 4 (index 3) = sub-headers
    concepts = []
    concept_rows = rows[4:] if len(rows) > 4 else []
    concept_rows = _merge_split_rows(concept_rows)

    for idx, row in enumerate(concept_rows):
        if _is_empty_row(row):
            continue
        # Skip summary / total rows (e.g. "Total Sessions")
        col1_val = _clean(row[1] if len(row) > 1 else None)
        if col1_val.lower().startswith("total"):
            continue

        s_no_raw = row[0] if row else None
        s_no = _clean(s_no_raw)
        if not s_no:
            # Science files have None in s_no column — fall back to row index
            s_no = str(idx + 1)

        content_title = _clean(row[1] if len(row) > 1 else None)
        sessions = _clean(row[2] if len(row) > 2 else None)
        exhibit_ref_raw = _clean(row[3] if len(row) > 3 else None)
        exhibit_ref = _normalize_exhibit_ref(exhibit_ref_raw) if exhibit_ref_raw else ""
        learning_outcomes = _clean(row[4] if len(row) > 4 else None)
        integration_other_sub = _clean(row[5] if len(row) > 5 else None)
        library = _clean(row[6] if len(row) > 6 else None)
        activity = _clean(row[7] if len(row) > 7 else None)

        # Remarks/life_lesson can be in col 8 or col 9 depending on subject
        remarks_val = _clean(row[8] if len(row) > 8 else None)
        life_lesson_val = _clean(row[9] if len(row) > 9 else None)

        concepts.append({
            "s_no": s_no,
            "title": content_title,
            "sessions": sessions,
            "exhibit_ref": exhibit_ref,
            "learning_outcomes": learning_outcomes,
            "integration_other_sub": integration_other_sub,
            "library": library,
            "activity": activity,
            "life_lesson": life_lesson_val,
            "remarks": remarks_val,
        })

    return title, aim, concepts


def _merge_split_rows(rows: list) -> list:
    """
    Some files split a single logical concept across multiple physical rows.
    The most common pattern: row N has s_no (col 0), row N+1 has no s_no but
    supplies sessions/exhibit_ref that were missing.

    Rules:
    - A row with an explicit s_no (col 0 non-None) always starts a new concept.
    - A row with NO s_no and NO title (col 1) is a continuation row — merge into pending.
    - A row with NO s_no but HAS a title:
        - If pending exists AND pending has NO exhibit_ref (col 3 None):
          merge into pending (this row fills in the exhibit_ref and sessions).
          Concatenate titles if both have title text.
        - Otherwise: this row stands alone as a new concept (e.g. Science where
          every row has no s_no but each has a distinct complete concept row).
    """
    if not rows:
        return []

    padded = []
    for row in rows:
        r = list(row)
        while len(r) < 10:
            r.append(None)
        padded.append(r)

    merged = []
    pending: list | None = None  # accumulated row being built

    for row in padded:
        if _is_empty_row(row):
            if pending is not None:
                merged.append(tuple(pending))
                pending = None
            continue

        has_sno = row[0] is not None
        has_title = row[1] is not None  # col 1 = content/title

        if has_sno:
            # Explicit s_no — flush previous and start a new concept
            if pending is not None:
                merged.append(tuple(pending))
            pending = row[:]
        elif not has_title:
            # Continuation row (no s_no, no title) — merge into pending if one
            # exists.  After an empty-row flush (pending is None) we skip these
            # orphan rows rather than letting them become a corrupt new pending.
            if pending is not None:
                for i in range(len(pending)):
                    if pending[i] is None and i < len(row) and row[i] is not None:
                        pending[i] = row[i]
            # else: pending was just flushed — discard this orphan continuation row
        else:
            # No s_no, has title
            if pending is not None and pending[3] is None:
                # Pending is missing exhibit_ref (col 3) — this row completes it
                for i in range(len(pending)):
                    if pending[i] is None and i < len(row) and row[i] is not None:
                        pending[i] = row[i]
                    elif i == 1 and pending[i] is not None and row[i] is not None:
                        # Concatenate titles
                        pending[i] = pending[i] + "\n" + row[i]
            else:
                # Pending has exhibit_ref OR no pending — flush and start new concept
                if pending is not None:
                    merged.append(tuple(pending))
                pending = row[:]

    if pending is not None:
        merged.append(tuple(pending))

    return merged


# ---------------------------------------------------------------------------
# Exhibit sheet parser
# ---------------------------------------------------------------------------

def _parse_exhibit_sheet(ws) -> dict:
    """
    Parse a single exhibit sheet.
    Returns {'raw_title': str, 'fields': {field_key: value, ...}}.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"raw_title": "", "fields": {}}

    # Determine raw title from the sheet
    first_row = rows[0]
    raw_title = _extract_exhibit_title(first_row)

    # Detect subject type from headers
    # Headers can be in row 1 (index 1) for Science/SST/Maths
    # or row 0 (index 0) for English/Hindi-style exhibits
    subject_type, header_row_idx, col_map = _detect_and_map_columns(rows, ws)

    fields = {}
    if col_map:
        data_start = header_row_idx + 1
        for row in rows[data_start:]:
            if _is_empty_row(row):
                continue
            for field_key, col_idx in col_map.items():
                if col_idx < len(row):
                    val = _clean(row[col_idx])
                    if val:
                        if field_key in fields:
                            fields[field_key] = fields[field_key] + "\n" + val
                        else:
                            fields[field_key] = val
    else:
        # Fallback for unknown/table-style exhibits: store raw text content
        # Skip only row 0 (title row); include from row 1 onward
        content_parts = []
        for row in rows[1:]:
            if _is_empty_row(row):
                continue
            row_text = " | ".join(_clean(v) for v in row if v is not None and _clean(v))
            if row_text:
                content_parts.append(row_text)
        if content_parts:
            fields["content"] = "\n".join(content_parts)

    # Ensure all expected field keys exist (no missing keys)
    for key in col_map:
        if key not in fields:
            fields[key] = ""

    return {"raw_title": raw_title, "fields": fields}


def _extract_exhibit_title(first_row: tuple) -> str:
    """Extract exhibit title from the first row of an exhibit sheet."""
    if not first_row:
        return ""
    # If col 0 is a string containing 'exhibit'/'exibit', use it
    val0 = first_row[0]
    if isinstance(val0, str):
        s = val0.lower()
        if "exhibit" in s or "exibit" in s:
            return _clean(val0)
    # If col 0 is a number (exhibit number) and col 1 is the title text
    if isinstance(val0, (int, float)) and len(first_row) > 1:
        col1 = first_row[1]
        if isinstance(col1, str) and col1.strip():
            return _clean(col1)
    # Scan cols 1-4 for a string containing 'exhibit'/'exibit'
    for v in list(first_row)[1:5]:
        if isinstance(v, str):
            sv = v.lower()
            if "exhibit" in sv or "exibit" in sv:
                return _clean(v)
    # Fallback: first non-None, non-empty string
    for v in first_row:
        if isinstance(v, str) and v.strip():
            return _clean(v)
    return ""


def _detect_and_map_columns(rows: list, ws=None) -> tuple:
    """
    Detect subject type and return (subject_type, header_row_idx, col_map).
    col_map maps field_key -> column index.
    ws is the openpyxl worksheet, used only for the warning message.
    """
    # Try row index 0 first (English-style exhibits where row 1 IS the header)
    # then row index 1 (Science/SST/Maths where row 2 is the header)
    for header_row_idx in [0, 1]:
        if header_row_idx >= len(rows):
            continue
        headers = [
            _clean(v).lower() if v is not None else "" for v in rows[header_row_idx]
        ]
        subject_type = _detect_subject_type(headers)
        if subject_type != "unknown":
            col_map = _build_col_map(subject_type, headers)
            return subject_type, header_row_idx, col_map

    # No recognized pattern — return unknown with empty col_map
    sheet_name = ws.title if ws is not None else "<unknown sheet>"
    print(f"Warning: unknown exhibit type in {sheet_name}, storing as raw content")
    return "unknown", 1, {}


def _detect_subject_type(headers: list) -> str:
    """
    Detect subject type from a list of cleaned lowercase header strings.
    Returns one of: 'english', 'grammar', 'assessment', 'science', 'unknown'.

    Priority order is intentional:
      1. assessment — checked first because "very short answer" headers are
         unambiguous and assessment sheets can appear inside English files
         (e.g. exhibit_3 of a poem file).
      2. grammar — checked before english because grammar exhibits also live
         inside English files and share some vocabulary; the grammar keywords
         are more specific and must win when present.
      3. english — broad "about the chapter / overview" keywords that would
         not appear in grammar or assessment sheets.
      4. science — catch-all for Science / SST / Maths exhibits that use
         "introductory / explanation / discussion" style headers.
    """
    # Assessment: "very short answer" type headers
    if any("very short" in h for h in headers):
        return "assessment"

    # Grammar: tongue twisters / word meaning / conjunction / adjective etc.
    grammar_keywords = [
        "conjunction", "adjective", "adverb", "tongue twister",
        "tongue twist", "word-meaning", "word meaning",
        "homophone", "grammar",
    ]
    if any(any(kw in h for kw in grammar_keywords) for h in headers):
        return "grammar"

    # English: about the chapter / poem, overview, life lesson
    english_keywords = [
        "about the chapter", "about the poem", "overview",
        "life lesson", "chapter overview", "poem overview",
        "ch. over view",
    ]
    if any(any(kw in h for kw in english_keywords) for h in headers):
        return "english"

    # Science/SST/Maths: introductory questions or explanation/discussion
    science_keywords = [
        "introduct", "introductory", "explanation",
        "discussion", "teaching method",
    ]
    if any(any(kw in h for kw in science_keywords) for h in headers):
        return "science"

    return "unknown"


def _build_col_map(subject_type: str, headers: list) -> dict:
    """Build field_key -> column_index mapping for the given subject type."""
    col_map = {}

    if subject_type == "english":
        for i, h in enumerate(headers):
            if "about" in h:
                col_map["about_chapter"] = i
            elif "overview" in h or "over view" in h:
                col_map["chapter_overview"] = i
            elif "life lesson" in h:
                col_map["life_lesson_detail"] = i
            elif "activity" in h:
                col_map["activity"] = i
            elif "link" in h:
                col_map["link"] = i

    elif subject_type == "grammar":
        # Match headers by keyword — same approach as english/assessment/science.
        # This is more robust than relying solely on column position, which
        # breaks when columns are reordered or a leading numeric/empty column
        # is absent.
        #
        # The grammar_topic column holds the name of the current grammar unit
        # (e.g. "Conjunction", "Adjective") and is NOT identified by a fixed
        # keyword.  Instead we assign it by position: the first non-numeric,
        # non-empty header that isn't matched by a more-specific keyword wins
        # as grammar_topic.  All other recognised columns (tongue twisters,
        # word meanings, exercises) are keyword-matched and take priority.
        for i, h in enumerate(headers):
            if not h or h.replace(".", "").strip().isdigit():
                continue
            if "tongue twister" in h or "tongue twist" in h:
                col_map["tongue_twisters"] = i
            elif "word meaning" in h or "word-meaning" in h:
                col_map["word_meanings"] = i
            elif "homophone" in h:
                # Homophones often share a column with word meanings; only
                # assign if word_meanings hasn't already been mapped.
                col_map.setdefault("word_meanings", i)
            elif "exercise" in h:
                col_map["grammar_exercises"] = i
            else:
                # Unrecognised text column — could be a grammar unit name
                # (e.g. "Conjunction") or a combined exercises column.
                # The first such column becomes grammar_topic; subsequent
                # ones become grammar_exercises if not yet assigned.
                if "grammar_topic" not in col_map:
                    col_map["grammar_topic"] = i
                else:
                    col_map.setdefault("grammar_exercises", i)

    elif subject_type == "assessment":
        for i, h in enumerate(headers):
            if "very short" in h:
                col_map["very_short_answers"] = i
            elif "short" in h and "very" not in h:
                col_map["short_answers"] = i
            elif "long" in h:
                col_map["long_answers"] = i
            elif "fact" in h or "observation" in h:
                col_map["fact_finding"] = i

    elif subject_type == "science":
        for i, h in enumerate(headers):
            if "introduct" in h:
                col_map["intro_questions"] = i
            elif "explanation" in h or "discussion" in h:
                col_map["explanation"] = i
            elif "teaching" in h or "method" in h:
                col_map["teaching_method"] = i
            elif "activity" in h and "teaching" not in h:
                col_map["teaching_method"] = col_map.get("teaching_method", i)
            elif "link" in h:
                col_map["link"] = i
            elif "assessment" in h or "remark" in h or "example" in h:
                col_map["assessment"] = i

    return col_map


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_exhibit_ref(ref: str) -> str:
    """
    Normalize exhibit ref strings to 'exhibit_N' format.
    Examples:
      'Exhibit 1'              -> 'exhibit_1'
      'exhibit 1'              -> 'exhibit_1'
      "'Exhibits 2'!B1"        -> 'exhibit_2'
      'Exhibit1'               -> 'exhibit_1'
      'exhibit 2'              -> 'exhibit_2'
      'Exhibit 3 ...'          -> 'exhibit_3'
      'See page 12, Exhibit 3' -> 'exhibit_3'  (not exhibit_12)
    Returns '' if no number found.

    Strategy: try to match the number that immediately follows the word
    "exhibit" (case-insensitive) first; fall back to the first digit sequence
    only when no "exhibit" keyword is present in the string.
    """
    if not ref:
        return ""
    ref_str = str(ref).strip()
    # Primary: anchor on the "exhibit" keyword to avoid false matches
    match = re.search(r'exhibi?ts?\s*(\d+)', ref_str, re.IGNORECASE)
    if match:
        return f"exhibit_{match.group(1)}"
    # Fallback: no "exhibit" keyword — grab first digit sequence
    match = re.search(r'\d+', ref_str)
    if match:
        return f"exhibit_{match.group()}"
    return ""


def _strip_aim_prefix(aim: str) -> str:
    """
    Remove 'Aim:', 'Aim-', 'Aim: -', 'Aim:-', 'Aim :' etc. prefixes.
    Handles compound prefixes like 'Aim: -' (colon + dash).
    """
    if not aim:
        return aim
    # Strip 'Aim' followed by any combination of whitespace, ':', '-'
    cleaned = re.sub(r'^Aim\s*:?\s*-?\s*', '', aim, flags=re.IGNORECASE).strip()
    return cleaned


def _clean(value) -> str:
    """Convert None to empty string and strip whitespace."""
    if value is None:
        return ""
    return str(value).strip()


def _is_empty_row(row) -> bool:
    """Return True if all cells in the row are None or empty strings."""
    return all(
        v is None or (isinstance(v, str) and v.strip() == "") for v in row
    )
